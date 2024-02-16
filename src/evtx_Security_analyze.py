import os
from collections import defaultdict
from Evtx.Evtx import Evtx
from xml.etree.ElementTree import fromstring
from datetime import datetime
import openpyxl

def extract_event_details(xml_content, namespaces):
    xml_root = fromstring(xml_content)
    details = {}

    # 基本事件資料段
    details['provider_name'] = xml_root.find('.//ns:Provider', namespaces).attrib.get('Name', '未找到')
    details['keywords'] = xml_root.find('.//ns:Keywords', namespaces).text
    details['level'] = xml_root.find('.//ns:Level', namespaces).text
    details['event_id'] = xml_root.find('.//ns:EventID', namespaces).text
    details['computer'] = xml_root.find('.//ns:Computer', namespaces).text
    details['channel'] = xml_root.find('.//ns:Channel', namespaces).text
    details['time_created'] = xml_root.find('.//ns:TimeCreated', namespaces).attrib.get('SystemTime', '未找到')

    # EventData 資料段
    event_data = {}
    for data in xml_root.findall('.//ns:EventData/ns:Data', namespaces):
        name = data.attrib.get('Name')
        if name:
            event_data[name] = data.text

    # 特定字段
    details.update({
        'target_username': event_data.get('TargetUserName', '未找到'),
        'target_domain_name': event_data.get('TargetDomainName', '未找到'),
        'logon_process_name': event_data.get('LogonProcessName', '未找到'),
        'ip_address': event_data.get('IpAddress', '未找到'),
        'ip_port': event_data.get('IpPort', '未找到')
    })

    return details

def audit_status(keywords):
    if keywords == '0x8010000000000000':
        return 'Fail'
    elif keywords == '0x8020000000000000':
        return 'Success'
    return 'Unknown'

def process_and_write_evtx(root_dir, output_xlsx, namespaces):
    wb = openpyxl.Workbook()
    
    # 初始化工作表
    sheet1 = wb.active
    sheet1.title = "Audit Failures"
    sheet1.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'TargetUserName', 'TargetDomainName', 'LogonProcessName', 'IpAddress', 'IpPort'])

    sheet2 = wb.create_sheet(title="EventID Count")
    sheet2.append(['EventID', 'Count'])

    sheet_4624 = wb.create_sheet(title="EventID 4624")
    sheet_4624.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'TargetUserName', 'TargetDomainName', 'LogonProcessName', 'IpAddress', 'IpPort'])

    sheet_4625 = wb.create_sheet(title="EventID 4625")
    sheet_4625.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'TargetUserName', 'TargetDomainName', 'LogonProcessName', 'IpAddress', 'IpPort'])

    sheet_4628 = wb.create_sheet(title="EventID 4648")
    sheet_4628.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'TargetUserName', 'TargetDomainName', 'IpAddress', 'IpPort'])

    event_id_counter = defaultdict(int)  # 初始化事件ID计数器

    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if file.startswith('Security_') and file.endswith('.evtx'):
                evtx_path = os.path.join(root, file)
                with Evtx(evtx_path) as evtx:
                    for record in evtx.records():
                        xml_content = record.xml()
                        details = extract_event_details(xml_content, namespaces)
                        details['audit_status'] = audit_status(details['keywords'])

                        event_id_counter[details['event_id']] += 1  # 统计 EventID

                        # Sheet1: Audit Failures
                        if details['audit_status'] == 'Fail':
                            sheet1.append([details['event_id'], details['provider_name'], details['level'], details['keywords'], details['audit_status'], details['computer'], details['channel'], details['time_created'], details['target_username'], details['target_domain_name'], details['logon_process_name'], details['ip_address'], details['ip_port']])

                        # Sheet_4624: EventID 4624
                        if details['event_id'] == '4624':
                            sheet_4624.append([details['event_id'], details['provider_name'], details['level'], details['keywords'], details['audit_status'], details['computer'], details['channel'], details['time_created'], details['target_username'], details['target_domain_name'], details['logon_process_name'], details['ip_address'], details['ip_port']])

                        # Sheet_4625: EventID 4625
                        if details['event_id'] == '4625':
                            sheet_4625.append([details['event_id'], details['provider_name'], details['level'], details['keywords'], details['audit_status'], details['computer'], details['channel'], details['time_created'], details['target_username'], details['target_domain_name'], details['logon_process_name'], details['ip_address'], details['ip_port']])

                        # Sheet_4648: EventID 4648
                        if details['event_id'] == '4648':
                            sheet_4628.append([details['event_id'], details['provider_name'], details['level'], details['keywords'], details['audit_status'], details['computer'], details['channel'], details['time_created'], details['target_username'], details['target_domain_name'], details['logon_process_name'], details['ip_address'], details['ip_port']])

    # 填充 Sheet2: EventID Count
    for event_id, count in event_id_counter.items():
        sheet2.append([event_id, count])

    wb.save(output_xlsx)

# 设置根目录路径和输出 XLSX 文件路径
root_dir = '202312/'
# output_xlsx = 'audit_data6_20240213.xlsx'

# Generate dynamic filename
now = datetime.now()
formatted_datetime = now.strftime("%Y%m%d%H%M%S")
output_xlsx = f"exportSecurity_{formatted_datetime}.xlsx" 
namespaces = {'ns': 'http://schemas.microsoft.com/win/2004/08/events/event'}

process_and_write_evtx(root_dir, output_xlsx, namespaces)
