import os
from collections import defaultdict
from Evtx.Evtx import Evtx
from xml.etree.ElementTree import fromstring
from datetime import datetime
import openpyxl


# 定義事件等級的中文
def translate_level(level):
    levels = {
        '4': '資訊',
        '3': '警告',
        '2': '錯誤',
        '1': '重大',
        '0': '資訊'
    }
    return levels.get(level, '其他')

def extract_event_details(xml_content, namespaces):
    xml_root = fromstring(xml_content)
    details = {}

    # 基本事件資料
    details['provider_name'] = xml_root.find('.//ns:Provider', namespaces).attrib.get('Name', '未找到')
    details['keywords'] = xml_root.find('.//ns:Keywords', namespaces).text
    details['level'] = translate_level(xml_root.find('.//ns:Level', namespaces).text)
    
    details['event_id'] = xml_root.find('.//ns:EventID', namespaces).text
    details['computer'] = xml_root.find('.//ns:Computer', namespaces).text
    details['channel'] = xml_root.find('.//ns:Channel', namespaces).text
    details['time_created'] = xml_root.find('.//ns:TimeCreated', namespaces).attrib.get('SystemTime', '未找到')

    # EventData 資料段
    data_values = [data.text for data in xml_root.findall('.//ns:EventData/ns:Data', namespaces)]
    details['EventData'] = ', '.join(filter(None, data_values))  # 過濾掉 None 並用逗號連接

    # EventData 資料段
    # event_data_values = []
    # for i, data in enumerate(xml_root.findall('.//ns:EventData/ns:Data', namespaces), start=1):
    #     event_data_values.append(data.text or '空')  # 使用 '空' 替代 None 或空字符串
    
    # # 組合 details 字典和 event_data_values
    # details.update({f'DataColumn{i}': value for i, value in enumerate(event_data_values, start=1)})

    return details

def audit_status(keywords):
    if keywords == '0x8010000000000000':
        return 'Fail'
    elif keywords == '0x8020000000000000':
        return 'Success'
    return 'Unknown'

def process_and_write_evtx(root_dir, output_xlsx, namespaces):
    wb = openpyxl.Workbook()
    
    # 初始化工作表
    sheet1 = wb.active
    sheet1.title = "Audit Failures"
    sheet1.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'EventData'])

    sheet2 = wb.create_sheet(title="EventID Count")
    sheet2.append(['EventID', 'Count'])

    sheet_level = wb.create_sheet(title="level")
    sheet_level.append(['EventID', 'ProviderName', 'Level', 'Keywords', 'AuditStatus', 'Computer', 'Channel', 'TimeCreated', 'EventData'])

    event_id_counter = defaultdict(int)  # 初始化事件ID

    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if file.startswith('Application_') and file.endswith('.evtx'):
                evtx_path = os.path.join(root, file)
                with Evtx(evtx_path) as evtx:
                    for record in evtx.records():
                        xml_content = record.xml()
                        details = extract_event_details(xml_content, namespaces)
                        details['audit_status'] = audit_status(details['keywords'])

                        # 統計 EventID
                        event_id_counter[details['event_id']] += 1

                        # Sheet1: Audit Failures
                        if details['audit_status'] == 'Fail':
                            sheet1.append([
                                details['event_id'], details['provider_name'], details['level'], 
                                details['keywords'], details['audit_status'], details['computer'], 
                                details['channel'], details['time_created'], details['EventData']])

                        # 0-資訊, 4-資訊,  3-警告, 2-錯誤, 1-重大
                        # 檢查 level 是否在目標列表中
                        if details['level'] in ['警告', '錯誤', '重大']:  
                            sheet_level.append([
                                details['event_id'], details['provider_name'], details['level'], 
                                details['keywords'], details['audit_status'], details['computer'], 
                                details['channel'], details['time_created'], details['EventData']])


    # Sheet2: EventID Count
    for event_id, count in event_id_counter.items():
        sheet2.append([event_id, count])

    wb.save(output_xlsx)

# 設定根目錄路徑和輸出 XLSX 文件路徑
root_dir = '202312/'
# output_xlsx = 'audit_data6_system20240214-7.xlsx'

# Generate dynamic filename
now = datetime.now()
formatted_datetime = now.strftime("%Y%m%d%H%M%S")
output_xlsx = f"exportApplycation_{formatted_datetime}.xlsx" 

namespaces = {'ns': 'http://schemas.microsoft.com/win/2004/08/events/event'}

process_and_write_evtx(root_dir, output_xlsx, namespaces)
